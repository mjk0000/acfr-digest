#!/usr/bin/env python3
"""
Independent identity audit of parse_cafr.py output workbooks.
Checks GAAP accounting identities and magnitude sanity — no reliance on the
parser's own logic or logs.
"""
import sys
from pathlib import Path
from openpyxl import load_workbook

NF = "NOT FOUND"
CATS = ['FB - Nonspendable', 'FB - Restricted', 'FB - Committed',
        'FB - Assigned', 'FB - Unassigned']

def num(v):
    return v if isinstance(v, (int, float)) else None

def audit_workbook(path, entity_kind_hint=''):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    idx = {h: i for i, h in enumerate(header)}
    findings = []
    stats = {'rows': 0, 'fb_identity_pass': 0, 'fb_identity_checkable': 0}

    for r in rows[1:]:
        if not r or not r[idx['Source File']]:
            continue
        stats['rows'] += 1
        src = r[idx['Source File']]
        g = lambda col: num(r[idx[col]])

        ta, tl = g('Total Assets'), g('Total Liabilities')
        tfb = g('Total Fund Balances')
        cats = [g(c) for c in CATS]
        rev, exp = g('Total Revenues'), g('Total Expenditures')
        ofs = g('Total OFS (Uses)')
        units = r[idx['Reporting Units']]
        rev_usd = num(r[idx['Total Revenues (USD)']]) if 'Total Revenues (USD)' in idx else None

        # 1. FB categories must sum to total fund balances
        if tfb is not None and all(c is not None for c in cats):
            stats['fb_identity_checkable'] += 1
            s = sum(cats)
            tol = max(2.0, abs(tfb) * 0.001)
            if abs(s - tfb) <= tol:
                stats['fb_identity_pass'] += 1
            else:
                findings.append(('FB-SUM', src,
                    f"categories sum {s:,.0f} != total fund balances {tfb:,.0f} "
                    f"(diff {s - tfb:+,.0f})"))

        # 2. Balance sheet structure: residual = assets - liabilities - FB
        #    equals net deferred inflows (>=0 expected; negative => suspect)
        if ta is not None and tl is not None and tfb is not None:
            resid = ta - tl - tfb
            # Small negative residual is LEGAL: rare entities (NC) report
            # deferred OUTFLOWS at fund level, making residual = deferred
            # inflows - deferred outflows, which can go negative. Only flag
            # residuals too large to plausibly be deferred outflows (>5% assets).
            if resid < -max(2.0, ta * 0.05):
                findings.append(('BS-RESID', src,
                    f"assets {ta:,.0f} < liabilities {tl:,.0f} + FB {tfb:,.0f} "
                    f"(residual {resid:,.0f}) — a captured value is likely wrong"))
            if tl < 0 or (tfb is not None and ta < 0):
                findings.append(('NEG', src, f"negative assets/liabilities: ta={ta} tl={tl}"))

        # 3. FB total must not exceed assets
        if ta is not None and tfb is not None and tfb > ta * 1.001:
            findings.append(('FB>ASSETS', src,
                f"fund balances {tfb:,.0f} exceed assets {ta:,.0f}"))

        # 4. RevEx plausibility
        if rev is not None and exp is not None and rev > 0 and exp > 0:
            ratio = rev / exp
            if not (0.25 <= ratio <= 4.0):
                findings.append(('REV/EXP', src,
                    f"revenues/expenditures ratio {ratio:.2f} out of range "
                    f"(rev {rev:,.0f}, exp {exp:,.0f})"))
        if ofs is not None and rev is not None and abs(ofs) > max(rev, exp or 0) * 1.05:
            findings.append(('OFS-MAG', src,
                f"|OFS| {abs(ofs):,.0f} exceeds revenues {rev:,.0f} — suspect"))

        # 5. Unit sanity: normalized GF revenues for any US city/county/state
        #    should be between $50M and $500B
        if rev_usd is not None and rev_usd > 0:
            if rev_usd < 5e7:
                findings.append(('UNIT-LOW', src,
                    f"normalized revenues ${rev_usd:,.0f} implausibly low "
                    f"(units='{units}') — unit detection suspect"))
            elif rev_usd > 5e11:
                findings.append(('UNIT-HIGH', src,
                    f"normalized revenues ${rev_usd:,.0f} implausibly high "
                    f"(units='{units}') — unit detection suspect"))

    return stats, findings

def main():
    base = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('.')
    total_findings = []
    for name in ['audit_set1.xlsx', 'audit_set2.xlsx', 'audit_set3.xlsx']:
        p = base / name
        if not p.exists():
            print(f"MISSING: {p}")
            continue
        stats, findings = audit_workbook(p)
        print(f"\n=== {name}: {stats['rows']} rows, FB identity "
              f"{stats['fb_identity_pass']}/{stats['fb_identity_checkable']} pass, "
              f"{len(findings)} findings")
        for tag, src, msg in findings:
            print(f"  [{tag}] {src.split('_ACFR')[0]}: {msg}")
        total_findings += findings
    print(f"\nTOTAL FINDINGS: {len(total_findings)}")

if __name__ == '__main__':
    main()
