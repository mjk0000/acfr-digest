"""Packaging smoke test for acfr-digest.

Runs the local Set2 corpus (5 city ACFRs — two of which are scanned PDFs and
known hard-fails, expected!) through parse_cafr.process_pdf and diffs every
output field against the audited baseline workbook from the 2026-07-07
release. Proves the packaging wave didn't change parser behavior. ~2 minutes.

Usage:  python3 tests/smoke.py
Exit 0 = green (all fields match the audited baseline), 1 = diff or setup error.
"""

import logging
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

from parse_cafr import NOT_FOUND, OUTPUT_COLUMNS, process_pdf  # noqa: E402

PDF_DIR = ROOT / 'FY25_City_ACFRs_Set2'
BASELINE = ROOT / 'Backups' / '2026-07-07_audited' / 'FY25_Set2_results.xlsx'

# write_excel strips control chars before writing strings; apply the same
# transform so note comparisons are like-for-like.
_CTRL = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')


def load_baseline():
    wb = load_workbook(BASELINE, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    col_idx = {name: header.index(name) for name in OUTPUT_COLUMNS}
    return {row[col_idx['Source File']]:
            {name: row[col_idx[name]] for name in OUTPUT_COLUMNS}
            for row in rows[1:]}


def normalize(value):
    if value is None:
        # An empty workbook cell reads back as None; process_pdf represents
        # "no notes" as ''. Numeric fields never produce None (NOT FOUND is
        # stored as its literal string).
        return ''
    if isinstance(value, str):
        return _CTRL.sub('', value)
    return float(value)


def main():
    if not PDF_DIR.is_dir() or not BASELINE.exists():
        print(f'SETUP ERROR: need {PDF_DIR} and {BASELINE} '
              '(local corpus + audited backup)')
        return 1

    baseline = load_baseline()
    logger = logging.getLogger('smoke')
    logger.addHandler(logging.NullHandler())

    failures = 0
    pdfs = sorted(PDF_DIR.glob('*.pdf'))
    print(f'Smoke test: {len(pdfs)} PDFs vs audited baseline {BASELINE.name}')
    for pdf_path in pdfs:
        expected = baseline.get(pdf_path.name)
        if expected is None:
            print(f'  FAIL {pdf_path.name}: not in baseline')
            failures += 1
            continue
        row = process_pdf(pdf_path, logger)
        diffs = []
        for col in OUTPUT_COLUMNS:
            got, want = normalize(row.get(col, NOT_FOUND)), normalize(expected[col])
            if got != want:
                diffs.append(f'{col}: got {got!r}, baseline {want!r}')
        if diffs:
            failures += 1
            print(f'  FAIL {pdf_path.name}:')
            for d in diffs:
                print(f'        {d}')
        else:
            print(f'  OK   {pdf_path.name}')

    if failures:
        print(f'\nSMOKE TEST FAILED: {failures} file(s) diverge from the '
              'audited baseline')
        return 1
    print('\nSMOKE TEST GREEN: all fields match the audited baseline')
    return 0


if __name__ == '__main__':
    sys.exit(main())
