"""
parse_cafr.py — ACFR General Fund Parser
=========================================
Parses U.S. municipal Annual Comprehensive Financial Reports (ACFRs/CAFRs)
and extracts General Fund figures from each report into a single Excel workbook.

Usage:
  python parse_cafr.py report_A.pdf report_B.pdf
  python parse_cafr.py --dir /path/to/reports/
  python parse_cafr.py report_A.pdf --dir /path/to/more/ --output results.xlsx --log parse.log

Requirements:
  Python 3.9+
  pip install -r requirements.txt
  brew install poppler   (provides pdftotext, pdffonts, pdftoppm)

Target figures extracted from the General Fund column only:
  Balance Sheet: Total Assets, Total Liabilities, FB-Nonspendable, FB-Restricted,
                 FB-Committed, FB-Assigned, FB-Unassigned, Total Fund Balances
  RevEx Statement: Total Revenues, Total Expenditures, Total OFS (Uses)
"""

import argparse
import logging
import re
import subprocess
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

NOT_FOUND = "NOT FOUND"

# Disqualifying terms for General Fund column identification (Rule 1)
# Each entry is a pattern that, if found in a normalized column name, disqualifies it.
DISQUALIFYING_PATTERNS = [
    re.compile(r'\bcapital\b'),
    re.compile(r'\bcip\b'),
    re.compile(r'\bimprovement\b'),
    re.compile(r'\bimprovements\b'),
    re.compile(r'\binfrastructure\b'),
    re.compile(r'\bconstruction\b'),
    re.compile(r'\bproject\b'),
    re.compile(r'\bprojects\b'),
    re.compile(r'\bdebt\b'),
    re.compile(r'\bbond\b'),
    re.compile(r'\bbonds\b'),
    re.compile(r'\bsinking\b'),
    re.compile(r'\borganic\b'),
    re.compile(r'\bgrant\b'),
    re.compile(r'\bgrants\b'),
    re.compile(r'\bwater\b'),
    re.compile(r'\bsewer\b'),
    re.compile(r'\bstormwater\b'),
    re.compile(r'\bparking\b'),
    re.compile(r'\btransit\b'),
    re.compile(r'\butility\b'),
    re.compile(r'\butilities\b'),
    re.compile(r'\benterprise\b'),
    re.compile(r'\bnon-major\b'),
    re.compile(r'\bnonmajor\b'),
    re.compile(r'\btotal\b'),
    re.compile(r'\btotals\b'),
    re.compile(r'\bcombining\b'),
    re.compile(r'\bcombined\b'),
    re.compile(r'\bactivities\b'),
    re.compile(r'\bspecial\b'),
    re.compile(r'\bservice\b'),  # "debt service" — but note: don't disqualify on service alone
                                  # if it's "General Services" this might be valid. However, any
                                  # fund named "[X] Service Fund" is typically non-operational.
                                  # Given the examples and GASB guidance, "service" in a fund name
                                  # (other than the General Fund) indicates a specific purpose.
]

# Unit detection patterns
UNIT_PATTERNS = {
    'thousands': re.compile(
        r'(?:expressed\s+in|in|amounts?\s+in|stated\s+in|reported\s+in|dollars?\s+in)'
        r'\s+thousands?'
        r'|000s?\s+omitted'
        r'|\(000s?\)'
        r'|\(?\$?000\'?s\)?'          # ($000's) / ($000s) / $000's / 000's — San Jose, DC styles
        r'|thousands?\s+of\s+dollars?'
        r'|in\s+thousands'
    ),
    'millions': re.compile(
        r'(?:expressed\s+in|in|amounts?\s+in|stated\s+in|reported\s+in|dollars?\s+in)'
        r'\s+millions?'
        r'|millions?\s+of\s+dollars?'
        r'|in\s+millions'
    ),
}

# Fund balance category detection patterns
FB_CATEGORY_PATTERNS = {
    'nonspendable': re.compile(r'\bnon.?spendable\b'),
    'restricted':   re.compile(r'\brestricted\b'),
    'committed':    re.compile(r'\bcommitted\b'),
    # 'unassigned' checked BEFORE 'assigned' — otherwise character-split labels
    # like "un assigned" match \bassigned\b first and get miscategorized.
    'unassigned':   re.compile(r'\bunassigned\b'),
    'assigned':     re.compile(r'\bassigned\b'),
}

# Stop conditions for Balance Sheet (after total fund balances, stop)
TOTAL_FB_PATTERN = re.compile(r'\btotal\s+fund\s+balances?\b')

# Patterns for RevEx structural rows
EXCESS_PATTERN = re.compile(r'\bexcess\b|\bdeficiency\b|\bdeficit\b')
NET_CHANGE_PATTERN = re.compile(r'\bnet\s+(change|increase|decrease)\s+in\s+fund\b')
OFS_SECTION_PATTERN = re.compile(r'\bother\s+financing\s+(sources?|uses?)\b|\btransfers?\s+(in|out)\b')
# Last alternative ('total ... other ... sources' without 'uses'/'financing') catches
# wrapped labels whose tail falls into the value band (Delaware: 'Total Other Sources
# (Uses) of / Financial Resources' — label zone sees only 'total other sources').
# Safe because the pattern is only consulted inside an open OFS section.
OFS_TOTAL_PATTERN = re.compile(r'\btotal\b.*\bfinancing\b|\btotal\b.*\bother\b.*\bfinancing\b|\btotal\b.*\bsources?\b.*\buses?\b|\btotal\b.*\bother\b.*\bsources?\b')

END_FB_PATTERN = re.compile(r'\bend\s+of\s+year\b|\bending\s+fund\b|\bend\s+of\s+period\b')

# Reconciliation stop signals (content that indicates reconciliation section)
RECONCILIATION_SIGNALS = [
    re.compile(r'\breconciliation\b'),
    re.compile(r'\bnot\s+financial\s+resources\b'),
    re.compile(r'\bcapital\s+assets\s+used\b'),
    re.compile(r'\blong.term\s+(liabilities|debt)\s+not\b'),
    re.compile(r'\bpension\b.*\badjustment\b'),
    re.compile(r'\bamounts\s+reported\s+for\s+governmental\s+activities\b'),
]

# Output columns in order
OUTPUT_COLUMNS = [
    'Source File', 'Entity Name', 'Fiscal Year End', 'Reporting Units',
    'Total Assets', 'Total Liabilities',
    'FB - Nonspendable', 'FB - Restricted', 'FB - Committed',
    'FB - Assigned', 'FB - Unassigned', 'Total Fund Balances',
    'Total Revenues', 'Total Expenditures', 'Total OFS (Uses)',
    'Extraction Notes',
]

# Columns that have a normalized USD counterpart (same order as OUTPUT_COLUMNS)
_NORMALIZABLE_COLS = [
    'Total Assets', 'Total Liabilities',
    'FB - Nonspendable', 'FB - Restricted', 'FB - Committed',
    'FB - Assigned', 'FB - Unassigned', 'Total Fund Balances',
    'Total Revenues', 'Total Expenditures', 'Total OFS (Uses)',
]

UNIT_MULTIPLIERS: Dict[str, int] = {
    'Thousands':   1_000,
    'Millions':    1_000_000,
    'Full Dollars': 1,
}

# ---------------------------------------------------------------------------
# Text utilities (Rule 6: Universal normalization)
# ---------------------------------------------------------------------------

def normalize(text: str) -> str:
    """
    Universal text normalization — apply before any label comparison.
    Handles whitespace artifacts, typographic characters, and case.
    """
    if not text:
        return ""
    # Typographic substitutions
    text = re.sub(r'[–—―‒]', '-', text)        # em/en dashes → hyphen
    text = re.sub(r'[''‘’]', "'", text)  # curly single quotes
    text = re.sub(r'["""“”]', '"', text)  # curly double quotes
    text = re.sub(r' ', ' ', text)          # non-breaking space
    # Collapse whitespace
    text = re.sub(r'\s+', ' ', text)
    return text.strip().lower()


def parse_number(text: str) -> Optional[float]:
    """
    Parse a formatted number string into a float.
    Handles: $ prefix, comma separators, parentheses for negative,
    em-dash / hyphen alone for zero.
    Returns None if the string cannot be parsed as a number.
    """
    if text is None:
        return None
    t = normalize(text).strip()
    # Bare dash or em-dash → zero (Rule 8)
    if t in ('-', '--', '—', '–', ''):
        return 0.0
    # Strip currency symbol
    t = t.replace('$', '').strip()
    if t in ('-', '--', ''):
        return 0.0
    # Parentheses → negative (Rule 8)
    negative = False
    if t.startswith('(') and t.endswith(')'):
        negative = True
        t = t[1:-1]
    # Trailing ) without matching ( is a column-separator artifact (e.g. Nashville)
    elif t.endswith(')') and not t.startswith('('):
        t = t[:-1].strip()
    # Strip commas and spaces
    t = t.replace(',', '').replace(' ', '')
    try:
        value = float(t)
        return -value if negative else value
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Unit detection (Core requirement: never skipped)
# ---------------------------------------------------------------------------

def detect_unit(page_text: str, prev_page_text: str = "") -> str:
    """
    Scan page text (and optionally preceding page) for unit disclosure.
    Returns 'Thousands', 'Millions', or 'Full Dollars'.
    """
    # Underscores stripped: decorative rule lines can interleave with the unit
    # text ('_in_ _T_h_o_u_s_a_n_d_s_' — Orange County).
    combined = normalize((prev_page_text or "") + " " + page_text).replace('_', '')
    # Compact form catches tight-kerning PDFs whose header renders '(InThousands)'
    # as one word (Broward) — the spaced patterns can't match those.
    compact = combined.replace(' ', '')
    if (UNIT_PATTERNS['millions'].search(combined)
            or 'inmillions' in compact or 'millionsofdollars' in compact):
        return 'Millions'
    if (UNIT_PATTERNS['thousands'].search(combined)
            or 'inthousands' in compact or 'thousandsofdollars' in compact
            or '000somitted' in compact):
        return 'Thousands'
    return 'Full Dollars'


# ---------------------------------------------------------------------------
# Page identification
# ---------------------------------------------------------------------------

def _get_title_area(page_text: str, chars: int = 900) -> str:
    """Return normalized first N chars of page text — where statement titles appear.
    Strips running 'table of contents' page headers printed on every page by some cities."""
    stripped = re.sub(r'(?i)^\s*table\s+of\s+contents\s*\n?', '', page_text)
    return normalize(stripped[:chars])


def is_balance_sheet_page(title_area: str, relaxed: bool = False) -> bool:
    """
    True if the page title area signals a Governmental Funds Balance Sheet (target).
    Checks concept first, excludes government-wide, combining, proprietary, etc.
    relaxed=True additionally accepts 'general fund' as the fund-statement
    signature (Tucson titles the page just "Balance Sheet") — SECOND-PASS ONLY:
    on a first pass it matches MD&A condensed tables that appear before the
    real statement (Idaho, Seattle) and steals the page slot.
    """
    has_bs = 'balance sheet' in title_area
    has_gf = ('governmental fund' in title_area or 'governmental balance' in title_area
              or (relaxed and 'general fund' in title_area))
    # Most bad words checked against full title_area (900 chars).
    bad = any(x in title_area for x in [
        'combining', 'proprietary', 'fiduciary',
        'table of contents', 'reconciliation',
        'net position',  # government-wide statement
    ])
    # MD&A narrative pages — prefix match tolerates in-document typos
    # (Orange County prints "Management's Disccusion and Analysis").
    bad = bad or bool(re.search(r"management'?s?\s+disc", title_area))
    # 'budget' scoped to the title zone only (first 250 chars): states commonly have
    # a "Budget Stabilization Fund" COLUMN HEADER on the real statement page, which
    # lands within 900 chars and must not disqualify the page.
    # 'notes to' guards against note-disclosure pages matching the relaxed
    # 'general fund' signature above.
    bad = bad or any(x in title_area[:250] for x in ('budget', 'notes to'))
    # 'component unit' checked only in the first 250 chars (statement title zone).
    # Philadelphia's BS has "due from component units" as a line-item label in the
    # data section (chars ~300-900); that must NOT exclude the page.
    # Exception: "(a component unit of ...)" is an entity self-description in the
    # running page header (e.g. Indianapolis Unigov structure) — not a section filter.
    if not bad:
        is_entity_descriptor = bool(re.search(r'\(a component unit\b', title_area[:250]))
        if not is_entity_descriptor:
            bad = bool(re.search(r'\bcomponent unit', title_area[:250]))
    return has_bs and has_gf and not bad


def is_revex_page(title_area: str, relaxed: bool = False) -> bool:
    """
    True if the page title area signals the Governmental Funds Revenue/Expenditure
    Statement (target). Excludes budget comparisons, combining schedules, etc.
    relaxed: see is_balance_sheet_page — second-pass fallback only.
    """
    has_rev = 'revenue' in title_area
    has_exp = 'expenditure' in title_area
    has_changes = 'changes in fund' in title_area
    has_gf = ('governmental fund' in title_area
              or (relaxed and 'general fund' in title_area))
    bad = any(x in title_area for x in [
        'combining', 'reconciliation',
        'proprietary', 'fiduciary', 'last ten', 'ten fiscal year',
        'table of contents',
        'expenses',  # proprietary funds use "expenses" not "expenditures"
        # 'comparative' removed — budget comparison schedules already caught by 'budget';
        # removing avoids false negatives on pages titled "...with comparative totals for YYYY"
    ])
    # MD&A — typo-tolerant prefix (see is_balance_sheet_page).
    bad = bad or bool(re.search(r"management'?s?\s+disc", title_area))
    # 'budget'/'notes to' scoped to title zone — see is_balance_sheet_page.
    bad = bad or any(x in title_area[:250] for x in ('budget', 'notes to'))
    return has_rev and has_exp and has_changes and has_gf and not bad


def text_layer_usable(pdf) -> bool:
    """
    Sample mid-document pages and check whether extracted text is real language
    or mojibake from broken font encodings (e.g. Arkansas: 'ÿ|23454ÿ789ÿ...',
    Atlanta: '/J7J;C;DJE<' glyph-offset text). Scanning such PDFs is pointless
    and extremely slow — bail out fast with a clear note instead.
    """
    n = len(pdf.pages)
    idxs = sorted({max(0, min(n - 1, i)) for i in (n // 4, n // 2, (3 * n) // 4)})
    scores = []
    for i in idxs:
        raw = (pdf.pages[i].extract_text() or "")[:2000]
        if len(raw) < 50:
            # Blank MID-document page: an ACFR's middle is dense text; emptiness
            # here means no usable text layer (image-only or dead encoding).
            scores.append(0.0)
            continue
        if raw.count('(cid:') >= 10:
            # pdfminer emits '(cid:NN)' placeholders for unmapped CID glyphs —
            # the letters in 'cid' would otherwise fool the letter-fraction test.
            scores.append(0.0)
            continue
        ascii_letters = sum(1 for c in raw if c.isalpha() and ord(c) < 128)
        scores.append(ascii_letters / len(raw))
    return (sum(scores) / len(scores)) >= 0.2


def find_statement_pages(pdf) -> Tuple[Optional[int], Optional[int]]:
    """
    Scan all pages to find the Balance Sheet and RevEx statement page indices (0-based).
    Returns (bs_page_idx, revex_page_idx). Either may be None if not found.
    """
    bs_idx = None
    revex_idx = None
    titles = {}  # cache: page idx -> title_area (for the relaxed second pass)

    for i, page in enumerate(pdf.pages):
        raw = page.extract_text() or ""
        title_area = _get_title_area(raw)
        titles[i] = title_area

        if bs_idx is None and is_balance_sheet_page(title_area):
            bs_idx = i
        if revex_idx is None and is_revex_page(title_area):
            revex_idx = i
        if bs_idx is not None and revex_idx is not None:
            break

    # Relaxed second pass — only for statements the strict pass missed.
    # Extends the title cache past the strict pass's early-break point.
    if bs_idx is None or revex_idx is None:
        for i in range(len(pdf.pages)):
            if i not in titles:
                titles[i] = _get_title_area(pdf.pages[i].extract_text() or "")
            if bs_idx is None and is_balance_sheet_page(titles[i], relaxed=True):
                bs_idx = i
            if revex_idx is None and is_revex_page(titles[i], relaxed=True):
                revex_idx = i
            if bs_idx is not None and revex_idx is not None:
                break

    return bs_idx, revex_idx


# ---------------------------------------------------------------------------
# Word/row utilities
# ---------------------------------------------------------------------------

def get_page_words(page) -> List[Dict]:
    """
    Extract words from a page using pdfplumber's character-level positional data.
    Uses x_tolerance=5 to handle minor spacing artifacts, y_tolerance=3 for row grouping.

    Tight-kerning PDFs (Allegheny) have inter-word gaps under 5pt, merging whole
    lines into single 'words' ('countyofallegheny,pennsylvania balancesheet').
    When long alphabetic run-ons dominate AND a tighter tolerance genuinely
    separates them, use the tighter extraction for that page.
    """
    words = page.extract_words(x_tolerance=5, y_tolerance=3, keep_blank_chars=False)
    if words:
        long_runs = sum(1 for w in words
                        if len(w['text']) > 22 and any(c.isalpha() for c in w['text']))
        if long_runs >= max(3, len(words) // 10):
            tighter = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
            if len(tighter) > len(words) * 1.5:
                return tighter
    return words


def cluster_into_rows(words: List[Dict], y_tolerance: float = 5.0) -> List[List[Dict]]:
    """
    Group words into rows by their vertical (y) center position.
    Returns list of rows, each row is a list of words sorted by x0.
    """
    if not words:
        return []
    rows: Dict[int, List[Dict]] = {}
    for w in words:
        y_center = (w['top'] + w['bottom']) / 2
        # Find existing row key within y_tolerance
        matched = None
        for yk in rows:
            if abs(y_center - yk) <= y_tolerance:
                matched = yk
                break
        if matched is None:
            matched = round(y_center)
            rows[matched] = []
        rows[matched].append(w)
    return [sorted(row_words, key=lambda w: w['x0']) for _, row_words in sorted(rows.items())]


def word_x_center(w: Dict) -> float:
    return (w['x0'] + w['x1']) / 2


# ---------------------------------------------------------------------------
# Column structure detection
# ---------------------------------------------------------------------------

def find_data_start_row(rows: List[List[Dict]], min_x: float = 180.0) -> int:
    """
    Return the index of the first row in `rows` that contains a financial
    figure at x > min_x. Requires a comma-separated number, $-prefix, or
    a parenthesis immediately followed by a digit (negative value) — this
    prevents false triggers on "(Expressed in thousands)", "(Deficits)", etc.
    Returns len(rows) if no financial figure is found (treat all rows as header).
    """
    # (?!\d) rejects date-like tokens: tight-kerning merges 'December31,2025'
    # into one token whose '31,202' would otherwise read as a financial number
    # and truncate the header zone above the real fund headers (Allegheny).
    data_num_pat = re.compile(r'^\$|^\(\d|\d{1,3},\d{3}(?!\d)')
    for i, row in enumerate(rows):
        if any(w['x0'] > min_x and data_num_pat.search(w['text']) for w in row):
            return i
    return len(rows)


_FILL_TOKEN = re.compile(r'^[_\-\.=\s]{2,}$')

def _is_fill_token(text: str) -> bool:
    """Return True for decorative fill words (underscores, dashes, dots).
    Catches both '____' style tokens AND 'f_u_nd_' style (letter-underscore interleaving)."""
    if _FILL_TOKEN.match(text):
        return True
    # Words with 2+ underscore characters are almost always fill decorations
    # (ACFR column headers never use underscores; these come from PDF underline rendering)
    return text.count('_') >= 2

def cluster_header_columns(header_words: List[Dict], x_gap: float = 30.0,
                           min_x: float = 180.0) -> List[Dict]:
    """
    Cluster header-zone words into column groups by x-position.
    Returns list of column dicts: {name, x_center, x_left, x_right, words}
    Ignores words whose x-center < min_x (row-label area) and fill tokens
    (runs of underscores/dashes/dots used as decorative lines in some PDFs).
    """
    # Filter out label-area words and fill-character tokens like "____" or "f_u_nd_"
    cand = [w for w in header_words
            if word_x_center(w) >= min_x and not _is_fill_token(w['text'])]
    if not cand:
        return []

    # Sort by x-center
    cand.sort(key=word_x_center)

    # Gap-based clustering
    clusters: List[List[Dict]] = []
    current: List[Dict] = [cand[0]]
    for w in cand[1:]:
        gap = word_x_center(w) - word_x_center(current[-1])
        if gap > x_gap:
            clusters.append(current)
            current = [w]
        else:
            current.append(w)
    clusters.append(current)

    # Build column descriptors
    col_list = []
    for cluster in clusters:
        xs = [word_x_center(w) for w in cluster]
        x_ctr = sum(xs) / len(xs)
        # Column name: join tokens ordered by y then x within y
        cluster_sorted = sorted(cluster, key=lambda w: (round(w['top'] / 6) * 6, w['x0']))
        name = normalize(' '.join(w['text'] for w in cluster_sorted))
        col_list.append({'name': name, 'x_center': x_ctr, 'words': cluster})

    # Set x_left and x_right boundaries as midpoints between adjacent columns
    for i, col in enumerate(col_list):
        if i == 0:
            col['x_left'] = max(0, col['x_center'] - 80)
        else:
            mid = (col_list[i - 1]['x_center'] + col['x_center']) / 2
            col['x_left'] = mid
            col_list[i - 1]['x_right'] = mid
        if i == len(col_list) - 1:
            col['x_right'] = col['x_center'] + 80

    return col_list


def identify_general_fund_col(columns: List[Dict]) -> Optional[Dict]:
    """
    Apply Rule 1 to identify the operational General Fund column.
    Returns the column dict or None if no valid column found.
    Handles character-split PDFs (e.g., "g en eral") by checking the space-removed
    compact form in addition to the original name.
    """
    # Disqualifying keywords for compact-form check (word-boundary regex can't apply
    # to space-stripped text, so we use substring matching instead).
    _DQ_KEYWORDS = [
        'capital', 'cip', 'improvement', 'infrastructure', 'construction', 'project',
        'debt', 'bond', 'sinking', 'grant', 'water', 'sewer', 'stormwater', 'parking',
        'transit', 'utility', 'enterprise', 'nonmajor', 'total', 'combining', 'combined',
        'activities', 'special', 'service',
    ]
    candidates = []
    for col in columns:
        name = col['name']
        compact = name.replace(' ', '')  # removes spaces for character-split detection

        # Must contain "general" in the spaced OR the compact form
        if 'general' not in name and 'general' not in compact:
            continue

        # Must not contain any disqualifying terms (check original with regex patterns)
        disqualified = any(pat.search(name) for pat in DISQUALIFYING_PATTERNS)
        # Also check compact form via substring (handles character-split disqualifiers)
        if not disqualified:
            disqualified = any(kw in compact for kw in _DQ_KEYWORDS)
        if disqualified:
            continue
        candidates.append(col)

    if not candidates:
        return None

    # Leftmost candidate is the General Fund (Rule 1, Step 3)
    return min(candidates, key=lambda c: c['x_center'])


def continuation_column_check(page, min_x: float = 180.0):
    """
    Light column-header detection for statement continuation pages (Fix A).
    Returns (columns, gf_col). columns is None when no header zone is detectable
    (pure vertical continuation — headers not repeated); gf_col is None when
    headers exist but none is the General Fund (horizontal continuation carrying
    OTHER funds at the same x-positions — caller must stop, not keep extracting).
    """
    words = get_page_words(page)
    rows = cluster_into_rows(words)
    data_start = find_data_start_row(rows, min_x=min_x)
    pre = [r for r in rows[:data_start] if r]
    if not pre:
        return None, None
    row_ys = [(r[0]['top'] + r[0]['bottom']) / 2 for r in pre]
    idx = 0
    for i in range(len(row_ys) - 1, 0, -1):
        if row_ys[i] - row_ys[i - 1] > 15.0:
            idx = i
            break
    header_words = [w for row in pre[idx:] for w in row]
    cols = cluster_header_columns(header_words, min_x=min_x)
    if not cols:
        return None, None
    return cols, identify_general_fund_col(cols)


def derive_label_gutter(rows: List[List[Dict]], default: float = 180.0) -> float:
    """
    Estimate where the row-label zone ends and data columns begin (Fix B).
    Uses the leftmost x0 among comma-formatted numbers, restricted to rows
    containing >=2 such numbers (real data rows span multiple columns; this
    excludes centered title dates like 'June 30, 2025').

    The derived boundary may only LOWER the default, never raise it: header
    words extend left of the values they sit above, so a boundary derived
    from value positions can overshoot and filter out real column headers
    (Sacramento: GF header at x=293 vs value-derived boundary of ~300).
    Dense statements that compress the gutter below 180 get the benefit;
    everything else behaves exactly as the historical 180pt constant.
    """
    _NUM = re.compile(r'^\$?\(?\d{1,3}(?:,\d{3})+\)?$')
    xs = []
    for row in rows:
        nums = [w for w in row if _NUM.match(w['text'].replace('_', ''))]
        if len(nums) >= 2:
            xs.extend(w['x0'] for w in nums)
    if not xs:
        return default
    return max(100.0, min(default, min(xs) - 12.0))


# ---------------------------------------------------------------------------
# Value extraction
# ---------------------------------------------------------------------------

def extract_value_in_column(row_words: List[Dict], col_left: float, col_right: float) -> Optional[float]:
    """
    Find and parse the numeric value within a column's x-range in a given row.
    Handles split numbers (e.g., Rockville "3 38,871" → 338,871),
    em-dash / hyphen zeros, and dot-leader fill characters (Baltimore "......" before values).
    """
    # +5pt right tolerance catches values printed just outside the column boundary
    # (e.g. SF where the GF column midpoint is 0.9pt inside the actual value position).
    tokens = [w for w in row_words if col_left <= word_x_center(w) <= col_right + 5]
    if not tokens:
        return None

    tokens.sort(key=lambda w: w['x0'])

    # Filter out non-numeric tokens: dot leaders ("......"), alphabetic stray words
    # that spill from narrow label zones into the column area (e.g., "balances", "inflows").
    # Keep tokens that: (a) contain a digit, OR (b) are pure numeric punctuation.
    # Strip underscores before testing — some PDFs (e.g., NYC) render digits with
    # underscore artifacts: '__5', '__,8', '__,' etc. After stripping, '5' has a digit
    # and ',' is pure punctuation.
    _NUMERIC_OR_PUNCT = re.compile(r'\d|^[\-–—\(\)\$,]+$')
    numeric = [t for t in tokens
               if _NUMERIC_OR_PUNCT.search(t['text'].replace('_', ''))]
    if not numeric:
        return None

    # Split-number fix: if the rightmost in-column token ends with a digit, check
    # whether the immediately adjacent token to the right starts with ',' — indicating
    # the number was split across the column boundary (e.g. SF RevEx: '6' | ',619,395').
    # Only check beyond the +5pt zone already collected above to avoid double-counting.
    rightmost = max(numeric, key=lambda w: w['x0'])
    if rightmost['text'].rstrip('_')[-1:].isdigit():
        continuation = [w for w in row_words
                        if col_right + 5 < word_x_center(w) < col_right + 20
                        and w['text'].lstrip('_').startswith(',')]
        if continuation:
            cont_tok = min(continuation, key=word_x_center)
            numeric.append(cont_tok)
            numeric.sort(key=lambda w: w['x0'])

    # Footnote refs ('Nonspendable (Note 1)') can bleed into the value band as
    # a '1)' token and poison the join ('1)6,485' → unparseable — Kentucky).
    # Drop them when a real comma-formatted number is also present in the band.
    if len(numeric) > 1:
        _REAL_NUM = re.compile(r'^\$?\(?\d{1,3}(?:,\d{3})+\)?$')
        if any(_REAL_NUM.match(t['text'].replace('_', '')) for t in numeric):
            numeric = [t for t in numeric if not re.match(r'^\d{1,2}\)$', t['text'])]

    # Sanity guard: two independently complete comma-formatted numbers in one band
    # means the band spans two real columns (merged headers on dense multi-fund
    # statements). Joining them fabricates an absurd value — refuse and return
    # None (NOT FOUND is the safe failure mode; a fabricated number is not).
    _FULL_NUM = re.compile(r'^\$?\(?\d{1,3}(?:,\d{3})+\)?$')
    complete_nums = [t for t in numeric if _FULL_NUM.match(t['text'].replace('_', ''))]
    if len(complete_nums) >= 2:
        return None

    combined = ''.join(t['text'] for t in numeric).replace('_', '').strip()

    # Em-dash or bare hyphen → zero
    if combined in ('-', '–', '—', '--'):
        return 0.0

    # Strip leading $ and try to parse
    result = parse_number(combined)
    return result


def row_label(row_words: List[Dict], col_left: float) -> str:
    """
    Extract and normalize the label portion of a row (words to the left of col_left).
    Also collapses character-split labels where individual letters appear isolated:
    e.g., "t otal assets" → "total assets", "n onspendable" → "nonspendable".
    """
    label_words = [w for w in row_words if word_x_center(w) < col_left - 5]
    raw = normalize(' '.join(w['text'] for w in sorted(label_words, key=lambda w: w['x0'])))
    # Collapse isolated single lowercase letters into the following token.
    # Only fires when the single char is NOT immediately preceded by another letter —
    # safe for normal labels while fixing "c ash" → "cash", "t otal" → "total".
    prev = None
    result = raw
    while result != prev:
        prev = result
        result = re.sub(r'(?<![a-z-])([a-z]) ([a-z])', r'\1\2', result)
    return result


# ---------------------------------------------------------------------------
# Balance Sheet extraction
# ---------------------------------------------------------------------------

def extract_bs_figures(
    pdf,
    start_page_idx: int,
    gf_col: Dict,
    notes: List[str],
    logger: logging.Logger,
) -> Dict[str, Any]:
    """
    Extract all Balance Sheet figures from the General Fund column.
    Handles multi-page statements and same-page reconciliation stop conditions.
    """
    results = {
        'total_assets': NOT_FOUND,
        'total_liabilities': NOT_FOUND,
        'nonspendable': NOT_FOUND,
        'restricted': NOT_FOUND,
        'committed': NOT_FOUND,
        'assigned': NOT_FOUND,
        'unassigned': NOT_FOUND,
        'total_fund_balances': NOT_FOUND,
    }

    col_left = gf_col['x_left']
    col_right = gf_col['x_right']

    # State machine for fund balance section
    in_fund_balance = False
    current_fb_cat = None        # category currently being accumulated
    fb_accumulator = 0.0         # running sum of sub-items
    fb_has_items = False         # whether any sub-items were seen
    fb_cat_found = set()         # categories whose header was found

    # Section tracking for unlabeled-total detection (Philadelphia pattern):
    # some PDFs omit "Total Assets" / "Total Liabilities" labels and instead
    # print the subtotal on a blank-label row after the section items.
    in_assets_section = False
    in_liabilities_section = False
    saw_asset_item = False
    saw_liability_item = False
    fallback_captured = set()    # keys captured via blank-label fallback (overwritable)
    prev_lbl = ''                # previous row's label (wrapped-header detection)
    prev_val = None              # previous row's value (wrapped-header detection)

    def finalize_fb_cat():
        nonlocal fb_accumulator, fb_has_items, current_fb_cat
        if current_fb_cat and current_fb_cat in fb_cat_found:
            if results[current_fb_cat] is NOT_FOUND and fb_has_items:
                results[current_fb_cat] = fb_accumulator
            elif results[current_fb_cat] is NOT_FOUND:
                results[current_fb_cat] = 0.0
        current_fb_cat = None
        fb_accumulator = 0.0
        fb_has_items = False

    stop_hit = False
    pages_checked = 0

    for page_idx in range(start_page_idx, min(start_page_idx + 3, len(pdf.pages))):
        if stop_hit:
            break
        if pages_checked > 0:
            # Continuation page: verify it doesn't start a NEW statement.
            # "(Continued)" pages repeat the BS title but are part of the same
            # statement — allow them through so fund balance data isn't missed.
            raw = pdf.pages[page_idx].extract_text() or ""
            title = _get_title_area(raw)
            is_continuation = 'continued' in title
            if not is_continuation and (is_balance_sheet_page(title) or is_revex_page(title)):
                break
            # Fix A — horizontal-continuation guard: re-detect column headers.
            # A repeated fund-header zone WITHOUT a General Fund column means this
            # page carries OTHER funds at the same x-positions; stop rather than
            # read another fund's numbers into the GF band. If the GF column
            # reappears (possibly re-typeset/shifted), adopt its band. No headers
            # at all = vertical continuation — keep page-1 band as before.
            cont_cols, cont_gf = continuation_column_check(pdf.pages[page_idx])
            if cont_gf is not None:
                col_left, col_right = cont_gf['x_left'], cont_gf['x_right']
            elif cont_cols is not None and len(cont_cols) >= 3:
                # Skip (not stop): 2x2 grid statements (Allegheny) alternate
                # GF-column pages with other-fund pages; the GF column returns
                # on the next page with the remaining rows.
                notes.append(f"BS p{page_idx + 1}: continuation carries other "
                             f"funds — page skipped")
                continue

        pages_checked += 1
        page = pdf.pages[page_idx]
        words = get_page_words(page)
        all_rows = cluster_into_rows(words)

        for row in all_rows:
            if not row:
                continue

            lbl = row_label(row, col_left)
            # Space-removed compact form handles character-split labels like
            # "fu nd ba la nc e" (fund balance) or "non spendable" (nonspendable)
            # where multi-char fragments prevent normal substring/regex matching.
            clbl = lbl.replace(' ', '')
            val = extract_value_in_column(row, col_left, col_right)

            # A 'fund balances' row directly after an unfinished wrapped header
            # ("Liabilities, Deferred Inflows of / ...and Fund Balances") is the
            # header's tail, not the section start — entering the FB section
            # there gates off the liabilities capture (Tucson). A prev row WITH
            # a value or starting with 'total' is a data/total row, not an
            # unfinished header ('Total deferred inflows of' precedes the real
            # FB header and must not suppress it).
            is_header_tail = (prev_lbl.endswith((' of', ' and', ',', ' inflows'))
                              and prev_val is None
                              and not prev_lbl.startswith('total'))
            prev_lbl, prev_val = lbl, val

            # Check for reconciliation signals (stop before reconciliation content)
            if any(sig.search(lbl) for sig in RECONCILIATION_SIGNALS):
                stop_hit = True
                finalize_fb_cat()
                break

            # --- Section tracking for blank-label totals (Philadelphia pattern) ---
            if re.match(r'^assets?:?$', lbl):
                in_assets_section = True
                in_liabilities_section = False
            elif re.match(r'^liabilit', lbl) and not in_fund_balance:
                in_liabilities_section = True
                in_assets_section = False

            # --- Balance Sheet top-level totals ---
            # Labeled totals may overwrite a fallback capture: blank-label group
            # subtotals (e.g. under 'Cash and investments:') can appear before the
            # real total row, and the fallback grabs them first (audit F2).
            if not in_fund_balance:
                if ('total assets' in lbl or 'totalassets' in clbl) and val is not None:
                    if results['total_assets'] is NOT_FOUND or 'total_assets' in fallback_captured:
                        results['total_assets'] = val
                        fallback_captured.discard('total_assets')
                # Fallback: blank-label row in assets section after items
                elif (results['total_assets'] is NOT_FOUND and in_assets_section
                      and saw_asset_item and lbl in ('', 'total') and val is not None):
                    results['total_assets'] = val
                    fallback_captured.add('total_assets')

            # Track individual asset items
            if in_assets_section and not in_liabilities_section and val is not None:
                if lbl not in ('', 'total') and not re.search(r'\btotal\b', lbl):
                    saw_asset_item = True

            if not in_fund_balance:
                if ('total liabilities' in lbl or 'totalliabilities' in clbl) and val is not None:
                    if results['total_liabilities'] is NOT_FOUND or 'total_liabilities' in fallback_captured:
                        results['total_liabilities'] = val
                        fallback_captured.discard('total_liabilities')
                # Fallback: blank-label row in liabilities section after items
                elif (results['total_liabilities'] is NOT_FOUND and in_liabilities_section
                      and saw_liability_item and lbl in ('', 'total') and val is not None):
                    results['total_liabilities'] = val
                    fallback_captured.add('total_liabilities')

            # Track individual liability items
            if in_liabilities_section and not in_fund_balance and val is not None:
                if lbl not in ('', 'total') and not re.search(r'\btotal\b', lbl):
                    saw_liability_item = True

            # --- Fund balance section entry ---
            # Exclude section header rows that contain "fund balance" but are NOT
            # the fund balance section entry: "LIABILITIES AND FUND BALANCES" (Houston,
            # Phoenix, NYC) and "RESOURCES AND FUND BALANCES" (Phoenix — split across
            # two rows so 'liabilit' doesn't appear on the second row).
            in_fb_lbl = re.search(r'\bfund\s+balance', lbl) or 'fundbalance' in clbl
            if (in_fb_lbl and 'total' not in lbl and 'totalfund' not in clbl
                    and 'liabilit' not in lbl and 'liabilit' not in clbl
                    and 'resources' not in lbl and not is_header_tail):
                in_fund_balance = True

            # --- Bare 'Total' closing an open category's sub-item list ---
            # While a category (e.g. 'Restricted for:') is accumulating sub-items,
            # an unlabeled 'Total' row is that category's subtotal — NOT total fund
            # balances. Capture it and keep scanning; treating it as the grand total
            # truncates the whole fund-balance section.
            if (in_fund_balance and lbl == 'total' and val is not None
                    and current_fb_cat is not None and fb_has_items):
                if results[current_fb_cat] is NOT_FOUND:
                    results[current_fb_cat] = val
                current_fb_cat = None
                fb_accumulator = 0.0
                fb_has_items = False
                continue

            # --- Total fund balances (stop condition, Rule 5) ---
            # Also catch "total fund" when 'balances' fell into the value zone (narrow label area).
            # Fallback: bare 'total' label in the fund balance section (Philadelphia pattern).
            total_fb_hit = (
                TOTAL_FB_PATTERN.search(lbl)
                or 'totalfundbalance' in clbl
                or ('total fund' in lbl and in_fund_balance)
                or ('totalfund' in clbl and in_fund_balance)
                or (in_fund_balance and lbl == 'total' and val is not None)
            )
            if total_fb_hit:
                finalize_fb_cat()
                if val is not None:
                    results['total_fund_balances'] = val
                stop_hit = True
                break

            if not in_fund_balance:
                continue

            # --- Fund balance category detection ---
            # Compact check uses startswith(cat) so "committedfor:" correctly
            # maps to 'committed' even though \bcommitted\b fails on merged text.
            matched_cat = None
            for cat, pat in FB_CATEGORY_PATTERNS.items():
                if pat.search(lbl) or clbl.startswith(cat):
                    matched_cat = cat
                    break

            if matched_cat and 'total' in lbl:
                # 'Total committed'-style subtotal row: close the open category
                # with the printed subtotal (more reliable than the accumulator);
                # never treat it as a new category row (would double-count).
                if current_fb_cat == matched_cat and val is not None:
                    if results[matched_cat] is NOT_FOUND:
                        results[matched_cat] = val
                    current_fb_cat = None
                    fb_accumulator = 0.0
                    fb_has_items = False
            elif matched_cat:
                # Finalize previous category's sub-items if any
                finalize_fb_cat()
                current_fb_cat = matched_cat
                fb_cat_found.add(matched_cat)
                fb_accumulator = 0.0
                fb_has_items = False

                # If the category row itself has a value → single-line format.
                # Repeated single-line rows for one category are SUMMED
                # ('Committed for X 5' / 'Committed for Y 7'); a later dash row
                # (0.0) must not wipe an earlier capture (Detroit: 'Committed
                # for community service —' overwrote Risk management 20,000,000).
                if val is not None:
                    if results[matched_cat] is NOT_FOUND:
                        results[matched_cat] = val
                    else:
                        results[matched_cat] += val
                    current_fb_cat = None  # Don't accumulate sub-items
            elif current_fb_cat is not None and results[current_fb_cat] is NOT_FOUND:
                # Sub-item row: accumulate
                if val is not None:
                    fb_accumulator += val
                    fb_has_items = True

    # If loop ended without hitting stop condition, finalize any open category
    if not stop_hit:
        finalize_fb_cat()

    return results


# ---------------------------------------------------------------------------
# Revenue/Expenditure Statement extraction
# ---------------------------------------------------------------------------

def extract_revex_figures(
    pdf,
    start_page_idx: int,
    gf_col: Dict,
    notes: List[str],
    logger: logging.Logger,
) -> Dict[str, Any]:
    """
    Extract Total Revenues, Total Expenditures, and Total OFS (Uses)
    from the General Fund column of the RevEx statement.
    Handles multi-page statements.
    """
    results = {
        'total_revenues': NOT_FOUND,
        'total_expenditures': NOT_FOUND,
        'total_ofs': NOT_FOUND,
    }

    col_left = gf_col['x_left']
    col_right = gf_col['x_right']

    in_ofs_section = False
    past_excess_line = False
    ofs_total_pending = False  # OFS total label found but value was on next row
    stop_hit = False
    pages_checked = 0
    in_revenues_section = False     # True once "revenues" header row is seen
    in_expenditures_section = False  # True once "expenditures" header row is seen
    saw_revenue_item = False        # True once at least one revenue line item found

    # For OFS: track the last "total" row in the OFS section as fallback
    last_total_in_ofs = None
    fallback_captured = set()    # keys captured via blank-label fallback (overwritable)

    for page_idx in range(start_page_idx, min(start_page_idx + 3, len(pdf.pages))):
        if stop_hit:
            break
        if pages_checked > 0:
            raw = pdf.pages[page_idx].extract_text() or ""
            title = _get_title_area(raw)
            is_continuation = 'continued' in title
            if not is_continuation and (is_balance_sheet_page(title) or is_revex_page(title)):
                break
            # Fix A — horizontal-continuation guard (see extract_bs_figures).
            cont_cols, cont_gf = continuation_column_check(pdf.pages[page_idx])
            if cont_gf is not None:
                col_left, col_right = cont_gf['x_left'], cont_gf['x_right']
            elif cont_cols is not None and len(cont_cols) >= 3:
                # Skip (not stop) — see extract_bs_figures (2x2 grid statements).
                notes.append(f"RevEx p{page_idx + 1}: continuation carries other "
                             f"funds — page skipped")
                ofs_total_pending = False
                continue
            # A pending two-row OFS capture must never satisfy across a page
            # break — the "next row" it expects no longer exists (audit F4).
            ofs_total_pending = False

        pages_checked += 1
        page = pdf.pages[page_idx]
        words = get_page_words(page)
        all_rows = cluster_into_rows(words)

        for row in all_rows:
            if not row:
                continue

            lbl = row_label(row, col_left)
            val = extract_value_in_column(row, col_left, col_right)

            # --- Two-row OFS label: label on one row, value on the next ---
            # Some PDFs split "Total other financing sources (uses)" across two rows.
            if ofs_total_pending:
                if val is not None and not NET_CHANGE_PATTERN.search(lbl):
                    results['total_ofs'] = val
                    in_ofs_section = False
                ofs_total_pending = False

            # --- Section header tracking (for unlabeled-total detection) ---
            if re.match(r'^revenues?:?$', lbl):
                in_revenues_section = True
                in_expenditures_section = False
            elif re.match(r'^expenditures?:?$', lbl):
                in_expenditures_section = True
                in_revenues_section = False

            # --- Total Revenues ---
            # Labeled total may overwrite a fallback capture (blank-label group
            # subtotals, e.g. under 'Taxes:', can precede the real total — audit F2).
            if re.search(r'\btotal\s+revenues?\b', lbl) and val is not None:
                if results['total_revenues'] is NOT_FOUND or 'total_revenues' in fallback_captured:
                    results['total_revenues'] = val
                    fallback_captured.discard('total_revenues')
            # Fallback: blank or bare "total" row in revenues section after items
            # (some PDFs print the subtotal without repeating "revenues" in the label)
            elif (results['total_revenues'] is NOT_FOUND and in_revenues_section
                  and saw_revenue_item and val is not None and lbl in ('', 'total')):
                results['total_revenues'] = val
                fallback_captured.add('total_revenues')

            # Track individual revenue items so we know when subtotal row arrives
            if in_revenues_section and val is not None and lbl not in ('', 'total') and not re.search(r'\btotal\b', lbl):
                saw_revenue_item = True

            # --- Total Expenditures ---
            if re.search(r'\btotal\s+expenditures?\b', lbl) and val is not None:
                if results['total_expenditures'] is NOT_FOUND or 'total_expenditures' in fallback_captured:
                    results['total_expenditures'] = val
                    fallback_captured.discard('total_expenditures')
            # Fallback: bare "total" in expenditures section
            elif (results['total_expenditures'] is NOT_FOUND and in_expenditures_section
                  and not past_excess_line and val is not None and lbl == 'total'):
                results['total_expenditures'] = val
                fallback_captured.add('total_expenditures')

            # --- Excess/Deficiency line (marks start of OFS section below it) ---
            if EXCESS_PATTERN.search(lbl) and 'total' not in lbl:
                past_excess_line = True

            # --- OFS section header detection ---
            # Allow OFS section to open if either: (a) we saw an excess/deficiency label,
            # or (b) total_expenditures is already captured (more robust than relying on
            # a specific label — e.g. Jacksonville has a typo "excesss" that breaks \bexcess\b).
            # Never re-open once total_ofs is captured: a later page's transfers rows
            # (another fund at the same x-positions) must not overwrite it (Johnson County).
            _past_exp = results['total_expenditures'] is not NOT_FOUND
            if ((past_excess_line or _past_exp) and OFS_SECTION_PATTERN.search(lbl)
                    and not in_ofs_section and results['total_ofs'] is NOT_FOUND):
                in_ofs_section = True

            # --- OFS total detection ---
            if in_ofs_section and OFS_TOTAL_PATTERN.search(lbl):
                if val is not None:
                    results['total_ofs'] = val
                    in_ofs_section = False
                else:
                    # Value might be on the next row (label wraps in the PDF)
                    ofs_total_pending = True
            elif in_ofs_section and re.match(r'^\(?uses?\)?:?$', lbl) and val is not None:
                # Charlotte-style split OFS total: the total label is split across rows
                # ('total other' with no value, then '(uses)' with the value). Capture
                # the '(uses)' row directly rather than using the pending mechanism,
                # which is too broad and causes cross-page false positives.
                results['total_ofs'] = val
                in_ofs_section = False

            # Track last "total" row in OFS section as fallback.
            # Exclude transfer subtotals ('Total transfers out') — they are line
            # items, not the section total, and the fallback must not promote them.
            if in_ofs_section and 'total' in lbl and 'transfer' not in lbl and val is not None:
                last_total_in_ofs = val

            # --- Net change in fund balance (stop condition) ---
            if NET_CHANGE_PATTERN.search(lbl):
                stop_hit = True
                break
            # Wrapped variant: deeply indented 'Net change in fund balances' can lose
            # its tail to the value band, leaving 'net change in' (Johnson County).
            # Gated on expenditures-found so revenue lines like 'net increase in
            # fair value of investments' can never trigger an early stop.
            if (_past_exp and re.search(r'\bnet\s+(changes?|increase|decrease)\s+in\b', lbl)):
                stop_hit = True
                break

            # --- End of year fund balance (also a stop condition) ---
            if END_FB_PATTERN.search(lbl):
                stop_hit = True
                break

    # If OFS total not found via label matching, try the last "total" in OFS section
    if results['total_ofs'] is NOT_FOUND and last_total_in_ofs is not None:
        results['total_ofs'] = last_total_in_ofs
        notes.append("OFS total found via positional fallback (label non-standard)")

    return results


# ---------------------------------------------------------------------------
# Column structure for a specific page
# ---------------------------------------------------------------------------

def get_column_structure(page, logger: logging.Logger) -> Optional[Dict]:
    """
    Identify the General Fund column on a given page.
    Returns column dict with x_left, x_right, x_center, name — or None.
    """
    words = get_page_words(page)
    all_rows = cluster_into_rows(words)

    # Fix B: derive the label/data boundary per page instead of assuming 180pt —
    # dense multi-fund state statements compress the label gutter; landscape
    # pages widen it. Falls back to 180 when the page has no qualifying numbers.
    gutter = derive_label_gutter(all_rows)

    # Use row-index-based separation so floating-point y comparisons can't
    # accidentally include the first data row in the header zone.
    data_start_idx = find_data_start_row(all_rows, min_x=gutter)
    pre_data_rows = all_rows[:data_start_idx]

    # Gap-based column header zone detection.
    # Strategy: try the LAST gap > 15px first (works for most PDFs). If that
    # yields no valid GF column, retry with the LARGEST gap (fixes PDFs like
    # LA where a section-header row sits just below the column names with a
    # small gap, eclipsing the larger title-to-column-header gap).
    row_ys = [(r[0]['top'] + r[0]['bottom']) / 2 for r in pre_data_rows if r]

    def _col_header_idx_last_gap(ys):
        for i in range(len(ys) - 1, 0, -1):
            if ys[i] - ys[i - 1] > 15.0:
                return i
        return 0

    def _col_header_idx_largest_gap(ys):
        best, best_i = 0.0, 0
        for i in range(1, len(ys)):
            gap = ys[i] - ys[i - 1]
            if gap > best and gap > 15.0:
                best, best_i = gap, i
        return best_i

    col_header_start_idx = _col_header_idx_last_gap(row_ys)
    header_words = [w for row in pre_data_rows[col_header_start_idx:] for w in row]
    columns = cluster_header_columns(header_words, min_x=gutter)
    gf_col = identify_general_fund_col(columns) if columns else None

    # Fallback 2: try largest-gap split when last-gap split finds nothing
    alt_idx = col_header_start_idx  # will be updated if fallback 2 runs
    if gf_col is None and row_ys:
        alt_idx = _col_header_idx_largest_gap(row_ys)
        if alt_idx != col_header_start_idx:
            alt_words = [w for row in pre_data_rows[alt_idx:] for w in row]
            alt_cols = cluster_header_columns(alt_words, min_x=gutter)
            alt_gf = identify_general_fund_col(alt_cols) if alt_cols else None
            if alt_gf:
                columns, gf_col = alt_cols, alt_gf

    # Fallback 3: backward scan — when both gap strategies pick a split point that
    # lands on label-zone-only rows (e.g. "Revenues:" at x=55), scan backward to
    # find the last split that produces valid candidate words at x >= 180.
    # Fixes: San Antonio RevEx, where the gap to "Revenues:" row slightly exceeds
    # the title-to-column-header gap.
    if gf_col is None and row_ys:
        tried = {col_header_start_idx, alt_idx}
        for try_idx in range(len(row_ys) - 2, -1, -1):
            if try_idx in tried:
                continue
            fb_words = [w for row in pre_data_rows[try_idx:] for w in row]
            fb_cand = [w for w in fb_words
                       if word_x_center(w) >= gutter and not _is_fill_token(w['text'])]
            if not fb_cand:
                continue
            fb_cols = cluster_header_columns(fb_words, min_x=gutter)
            fb_gf = identify_general_fund_col(fb_cols) if fb_cols else None
            if fb_gf:
                columns, gf_col = fb_cols, fb_gf
                break

    # Fallback 4: dense statements can merge adjacent headers into one cluster
    # (e.g. Mecklenburg: 'community development general fund block grants',
    # which the 'grant' keyword then disqualifies). Retry with tighter gaps —
    # only reached when every normal pass failed, so zero regression surface.
    if gf_col is None and header_words:
        for tight_gap in (18.0, 12.0):
            t_cols = cluster_header_columns(header_words, x_gap=tight_gap, min_x=gutter)
            t_gf = identify_general_fund_col(t_cols) if t_cols else None
            if t_gf:
                columns, gf_col = t_cols, t_gf
                logger.info(f"General Fund column found via tight header "
                            f"clustering (x_gap={tight_gap})")
                break

    if not columns:
        logger.warning("No column structure found on page")
        return None

    if gf_col is None:
        col_names = [c['name'] for c in columns]
        logger.warning(f"No General Fund column identified. Columns found: {col_names}")
        return None

    logger.info(f"General Fund column identified: '{gf_col['name']}' "
                f"(x_center={gf_col['x_center']:.1f}, "
                f"x_range=[{gf_col['x_left']:.1f}, {gf_col['x_right']:.1f}])")

    # Leftmost-position cross-check (Rule 1, Step 3)
    # Exclude long-name entries — page title text that leaked into column clustering
    # has a sentence-length name; real column headers are always short.
    non_total_cols = [c for c in columns
                      if 'total' not in c['name'] and 'totals' not in c['name']
                      and len(c['name']) <= 35]
    if non_total_cols:
        leftmost = min(non_total_cols, key=lambda c: c['x_center'])
        if abs(leftmost['x_center'] - gf_col['x_center']) > 20:
            logger.warning(
                f"General Fund column (x={gf_col['x_center']:.1f}) is NOT the leftmost "
                f"data column (leftmost: '{leftmost['name']}' x={leftmost['x_center']:.1f}). "
                f"Proceeding with conceptual identification — verify manually."
            )

    return gf_col


# ---------------------------------------------------------------------------
# pdftotext fallback (Layer 2)
# ---------------------------------------------------------------------------

def pdftotext_available() -> bool:
    try:
        subprocess.run(['pdftotext', '-v'], capture_output=True, timeout=5)
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def extract_via_pdftotext(pdf_path: Path, page_num: int) -> str:
    """Run pdftotext -layout on a single page and return the text."""
    result = subprocess.run(
        ['pdftotext', '-layout', '-f', str(page_num), '-l', str(page_num),
         str(pdf_path), '-'],
        capture_output=True, text=True, timeout=30
    )
    return result.stdout


# ---------------------------------------------------------------------------
# OCR fallback (Layer 3)
# ---------------------------------------------------------------------------

def is_scanned_pdf(pdf_path: Path) -> bool:
    """Check if PDF has no embedded fonts (scanned/raster)."""
    try:
        result = subprocess.run(
            ['pdffonts', str(pdf_path)], capture_output=True, text=True, timeout=30
        )
        lines = result.stdout.strip().splitlines()
        # pdffonts output: 2-line header + one line per font. No fonts = scanned.
        return len(lines) <= 2
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def ocr_page(pdf_path: Path, page_num: int) -> str:
    """Rasterize a page at 300 DPI and run pytesseract OCR."""
    try:
        import pytesseract
        from PIL import Image
        import tempfile, os

        with tempfile.TemporaryDirectory() as tmpdir:
            prefix = os.path.join(tmpdir, 'page')
            subprocess.run(
                ['pdftoppm', '-r', '300', '-f', str(page_num), '-l', str(page_num),
                 '-png', str(pdf_path), prefix],
                capture_output=True, timeout=60, check=True
            )
            images = sorted(Path(tmpdir).glob('*.png'))
            if not images:
                return ""
            img = Image.open(images[0])
            return pytesseract.image_to_string(img)
    except Exception as e:
        return ""


# ---------------------------------------------------------------------------
# Entity name and fiscal year extraction
# ---------------------------------------------------------------------------

def extract_entity_info(pdf) -> Tuple[str, str]:
    """
    Extract municipality name and fiscal year end from the first ~15 pages.
    Large city ACFRs often have lengthy front matter before the cover page text.
    """
    entity_name = NOT_FOUND
    fy_end = NOT_FOUND

    # "June 30, 2025" / "June 30 2025"
    date_pat = re.compile(
        r'(january|february|march|april|may|june|july|august|september|'
        r'october|november|december)\s+\d{1,2},?\s+\d{4}',
        re.IGNORECASE
    )
    # "fiscal year ended June 30, 2025" / "year ended September 30, 2024"
    year_ended_pat = re.compile(
        r'(?:fiscal\s+year|year)\s+ended?\s+'
        r'((?:january|february|march|april|may|june|july|august|september|'
        r'october|november|december)\s+\d{1,2},?\s+\d{4})',
        re.IGNORECASE
    )
    # "City/County/Town of X" — stops at comma, newline, or state name indicators
    city_pat = re.compile(
        r'(?:city|county|town|village|township|borough|municipality)\s+of\s+'
        r'([a-z][a-z\s\-\.]+?)(?:\s*,|\s*\n|$)',
        re.IGNORECASE
    )
    # State/commonwealth pattern is CASE-SENSITIVE: "state of" is common English
    # prose ("the state of the economy" in a Governor's transmittal letter would
    # match an IGNORECASE version). Cover pages always capitalize: "State of Alaska",
    # "STATE OF ALASKA", "Commonwealth of Virginia", "District of Columbia".
    state_pat = re.compile(
        r'(?:STATE|State|COMMONWEALTH|Commonwealth|DISTRICT|District)\s+'
        r'(?:OF|of|Of)\s+([A-Z][A-Za-z\s\-\.]+?)(?:\s*,|\s*\n|$)'
    )

    pages_to_check = min(15, len(pdf.pages))

    for i in range(pages_to_check):
        page = pdf.pages[i]
        raw = page.extract_text() or ""
        words = get_page_words(page)
        word_text = ' '.join(w['text'] for w in words)

        # Fiscal year end — prefer "year ended X" phrasing, fall back to bare date
        if fy_end is NOT_FOUND:
            for text_src in [raw, word_text]:
                m = year_ended_pat.search(text_src)
                if m:
                    fy_str = re.sub(r'\s+', ' ', m.group(1).strip())
                    fy_end = re.sub(
                        r'^(january|february|march|april|may|june|july|august|'
                        r'september|october|november|december)',
                        lambda x: x.group().capitalize(),
                        fy_str, flags=re.IGNORECASE
                    )
                    break
            if fy_end is NOT_FOUND:
                for text_src in [raw, word_text]:
                    m = date_pat.search(text_src)
                    if m:
                        fy_str = re.sub(r'\s+', ' ', m.group(0).strip())
                        fy_end = re.sub(
                            r'^(january|february|march|april|may|june|july|august|'
                            r'september|october|november|december)',
                            lambda x: x.group().capitalize(),
                            fy_str, flags=re.IGNORECASE
                        )
                        break

        # Entity name
        if entity_name is NOT_FOUND:
            for text_src in [raw, word_text]:
                m = city_pat.search(text_src) or state_pat.search(text_src)
                if m:
                    entity_name = m.group(0).strip().rstrip(',').rstrip()
                    entity_name = re.sub(r'\s+', ' ', entity_name)
                    break

            # Fallback: all-caps lines containing a govt-unit keyword
            if entity_name is NOT_FOUND:
                for line in raw.splitlines()[:10]:
                    line = line.strip()
                    if re.match(r'^[A-Z][A-Z\s,\.\']+$', line) and 5 < len(line) < 80:
                        if any(kw in line.lower() for kw in
                               ['city', 'county', 'town', 'village', 'township',
                                'borough', 'municipality', 'parish', 'district',
                                'state', 'commonwealth']):
                            entity_name = line.title()
                            break

        if entity_name is not NOT_FOUND and fy_end is not NOT_FOUND:
            break

    # Clean up entity name
    if entity_name is not NOT_FOUND:
        # Strip report-type suffixes that sometimes get captured with the name
        _REPORT_SUFFIXES = re.compile(
            r'\s*(?:comprehensive\s+)?(?:annual\s+)?(?:comprehensive\s+)?financial\s+report.*$'
            r'|\s*cafr.*$|\s*acfr.*$',
            re.IGNORECASE
        )
        entity_name = _REPORT_SUFFIXES.sub('', entity_name).strip().rstrip(',').strip()
        # Title-case all-caps names (CITY OF NEW YORK → City of New York)
        if entity_name == entity_name.upper():
            entity_name = entity_name.title()
        # Title-casing capitalizes connectives: "State Of Alaska" → "State of Alaska"
        entity_name = re.sub(r'\bOf\b', 'of', entity_name)

    return entity_name, fy_end


# ---------------------------------------------------------------------------
# Main PDF processing
# ---------------------------------------------------------------------------

def process_pdf(pdf_path: Path, logger: logging.Logger) -> Dict[str, Any]:
    """
    Process a single ACFR PDF and return a result dict for one Excel row.
    """
    notes: List[str] = []
    result: Dict[str, Any] = {col: NOT_FOUND for col in OUTPUT_COLUMNS}
    result['Source File'] = pdf_path.name

    logger.info(f"\n{'='*60}")
    logger.info(f"Processing: {pdf_path.name}")
    logger.info(f"{'='*60}")

    # --- Layer 3 check: detect scanned PDFs ---
    scanned = is_scanned_pdf(pdf_path)
    if scanned:
        notes.append("No embedded fonts detected (scanned or corrupt PDF) — text extraction may fail")
        logger.warning(f"{pdf_path.name}: No embedded fonts detected — scanned or corrupt PDF")

    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            logger.info(f"Total pages: {total_pages}")

            # --- Mojibake fast-bail: broken font encodings make every page
            # extract as garbage; a full scan is pointless and can cost 30+ min.
            if not text_layer_usable(pdf):
                notes.append("Text layer unusable (broken font encoding) — "
                             "needs OCR; skipped full scan")
                logger.error(f"{pdf_path.name}: text layer unusable (mojibake) — skipping")
                result['Extraction Notes'] = ' | '.join(notes)
                return result

            # --- Extract entity info ---
            entity_name, fy_end = extract_entity_info(pdf)
            result['Entity Name'] = entity_name
            result['Fiscal Year End'] = fy_end
            logger.info(f"Entity: {entity_name}, FY End: {fy_end}")

            # --- Find target statement pages ---
            bs_page_idx, revex_page_idx = find_statement_pages(pdf)
            logger.info(f"Balance Sheet page: {bs_page_idx + 1 if bs_page_idx is not None else 'NOT FOUND'}")
            logger.info(f"RevEx page: {revex_page_idx + 1 if revex_page_idx is not None else 'NOT FOUND'}")

            if bs_page_idx is None and revex_page_idx is None:
                notes.append("Could not locate either target statement — manual review required")
                logger.error(f"{pdf_path.name}: Neither target statement found")
                result['Extraction Notes'] = ' | '.join(notes)
                return result

            # --- Positional rescue: one statement found, the other's title
            # unreadable (e.g. San Antonio's BS title font has a broken cmap).
            # GASB layout places BS a few pages before RevEx; scan the
            # neighborhood for the nearest page with a General Fund column.
            if bs_page_idx is None and revex_page_idx is not None:
                for cand in range(revex_page_idx - 1, max(-1, revex_page_idx - 5), -1):
                    if get_column_structure(pdf.pages[cand], logging.getLogger('probe')):
                        bs_page_idx = cand
                        notes.append(f"BS page {cand + 1} inferred by position "
                                     f"(title unreadable) — verify manually")
                        logger.info(f"Balance Sheet page {cand + 1} inferred by position")
                        break
            elif revex_page_idx is None and bs_page_idx is not None:
                for cand in range(bs_page_idx + 1, min(len(pdf.pages), bs_page_idx + 5)):
                    if get_column_structure(pdf.pages[cand], logging.getLogger('probe')):
                        revex_page_idx = cand
                        notes.append(f"RevEx page {cand + 1} inferred by position "
                                     f"(title unreadable) — verify manually")
                        logger.info(f"RevEx page {cand + 1} inferred by position")
                        break

            # --- Reporting Unit Detection ---
            # Detect per statement page; flag disagreement rather than silently
            # applying one page's unit to the other statement's figures.
            unit = 'Full Dollars'
            page_units = {}
            for idx in [bs_page_idx, revex_page_idx]:
                if idx is not None:
                    page_text = pdf.pages[idx].extract_text() or ""
                    prev_text = pdf.pages[idx - 1].extract_text() if idx > 0 else ""
                    page_units[idx] = detect_unit(page_text, prev_text)
            detected_units = {u for u in page_units.values() if u != 'Full Dollars'}
            if detected_units:
                unit = sorted(detected_units)[0]
            if len(set(page_units.values())) > 1:
                notes.append(f"Unit disclosure differs between statement pages: "
                             f"{page_units} — verify normalized figures manually")
                logger.warning(f"Unit disagreement across statement pages: {page_units}")
            if unit == 'Full Dollars':
                logger.info("No unit notation detected — defaulting to Full Dollars")
            else:
                logger.info(f"Reporting units detected: {unit}")
            result['Reporting Units'] = unit

            # ================================================================
            # Balance Sheet extraction
            # ================================================================
            if bs_page_idx is not None:
                bs_page = pdf.pages[bs_page_idx]
                gf_col = get_column_structure(bs_page, logger)

                # Two-page-spread rescue (Orange County): the statement TITLE is
                # printed on the RIGHT page of a facing-page spread while the
                # General Fund column sits on the LEFT page (which has no title
                # of its own). If the title page has no GF column, try the page
                # before it.
                if gf_col is None and bs_page_idx > 0:
                    prev_gf = get_column_structure(pdf.pages[bs_page_idx - 1], logger)
                    if prev_gf is not None:
                        bs_page_idx -= 1
                        gf_col = prev_gf
                        notes.append(f"BS: spread layout — General Fund column on "
                                     f"facing page {bs_page_idx + 1}")
                        logger.info(f"BS spread layout: using facing page {bs_page_idx + 1}")

                if gf_col is None:
                    notes.append(f"BS p{bs_page_idx + 1}: General Fund column not identified")
                    logger.warning(f"BS: Could not identify General Fund column")
                else:
                    bs_figures = extract_bs_figures(pdf, bs_page_idx, gf_col, notes, logger)
                    result['Total Assets'] = bs_figures['total_assets']
                    result['Total Liabilities'] = bs_figures['total_liabilities']
                    result['FB - Nonspendable'] = bs_figures['nonspendable']
                    result['FB - Restricted'] = bs_figures['restricted']
                    result['FB - Committed'] = bs_figures['committed']
                    result['FB - Assigned'] = bs_figures['assigned']
                    result['FB - Unassigned'] = bs_figures['unassigned']
                    result['Total Fund Balances'] = bs_figures['total_fund_balances']

                    # Log NOT FOUND items
                    for k, v in bs_figures.items():
                        if v is NOT_FOUND:
                            notes.append(f"NOT FOUND: {k}")
                            logger.warning(f"BS: {k} = NOT FOUND")
                        else:
                            logger.info(f"BS: {k} = {v:,.2f}")
            else:
                notes.append("Balance Sheet page not found")

            # ================================================================
            # RevEx extraction
            # ================================================================
            if revex_page_idx is not None:
                revex_page = pdf.pages[revex_page_idx]

                # The RevEx statement shares the same multi-column structure as the BS.
                # Re-detect column structure on this page.
                gf_col_revex = get_column_structure(revex_page, logger)

                # Two-page-spread rescue — see Balance Sheet block above.
                if gf_col_revex is None and revex_page_idx > 0:
                    prev_gf = get_column_structure(pdf.pages[revex_page_idx - 1], logger)
                    if prev_gf is not None:
                        revex_page_idx -= 1
                        gf_col_revex = prev_gf
                        notes.append(f"RevEx: spread layout — General Fund column on "
                                     f"facing page {revex_page_idx + 1}")
                        logger.info(f"RevEx spread layout: using facing page {revex_page_idx + 1}")

                if gf_col_revex is None:
                    notes.append(f"RevEx p{revex_page_idx + 1}: General Fund column not identified")
                    logger.warning("RevEx: Could not identify General Fund column")
                else:
                    revex_figures = extract_revex_figures(
                        pdf, revex_page_idx, gf_col_revex, notes, logger
                    )
                    result['Total Revenues'] = revex_figures['total_revenues']
                    result['Total Expenditures'] = revex_figures['total_expenditures']
                    result['Total OFS (Uses)'] = revex_figures['total_ofs']

                    for k, v in revex_figures.items():
                        if v is NOT_FOUND:
                            notes.append(f"NOT FOUND: {k}")
                            logger.warning(f"RevEx: {k} = NOT FOUND")
                        else:
                            logger.info(f"RevEx: {k} = {v:,.2f}")
            else:
                notes.append("RevEx statement page not found")

    except Exception as e:
        logger.exception(f"Unhandled error processing {pdf_path.name}: {e}")
        notes.append(f"Processing error: {e}")

    result['Extraction Notes'] = ' | '.join(notes) if notes else ''
    return result


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_excel(all_results: List[Dict], output_path: Path, logger: logging.Logger):
    """Write all results to an Excel workbook per the output specification."""
    wb = Workbook()
    ws = wb.active
    ws.title = "General Fund Data"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(start_color="2B5496", end_color="2B5496", fill_type="solid")
    header_fill_usd = PatternFill(start_color="1F6B4E", end_color="1F6B4E", fill_type="solid")
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Standard column headers (blue)
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # USD normalized column headers (green), appended after standard columns
    usd_col_start = len(OUTPUT_COLUMNS) + 1
    for i, src_col in enumerate(_NORMALIZABLE_COLS):
        cell = ws.cell(row=1, column=usd_col_start + i, value=f'{src_col} (USD)')
        cell.font = header_font
        cell.fill = header_fill_usd
        cell.alignment = header_align

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    numeric_col_names = set(_NORMALIZABLE_COLS)
    number_format = '#,##0.00_);(#,##0.00)'

    # Data rows
    for row_idx, row_data in enumerate(all_results, 2):
        # Standard columns
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, 1):
            value = row_data.get(col_name, NOT_FOUND)
            cell  = ws.cell(row=row_idx, column=col_idx)
            if col_name in numeric_col_names and value is not NOT_FOUND:
                cell.value        = value
                cell.number_format = number_format
                cell.alignment    = Alignment(horizontal='right')
            else:
                cell.value     = value
                cell.alignment = Alignment(horizontal='left',
                                           wrap_text=(col_name == 'Extraction Notes'))

        # USD normalized columns
        unit_str   = row_data.get('Reporting Units') or 'Full Dollars'
        multiplier = UNIT_MULTIPLIERS.get(unit_str, 1)
        for i, src_col in enumerate(_NORMALIZABLE_COLS):
            raw  = row_data.get(src_col, NOT_FOUND)
            cell = ws.cell(row=row_idx, column=usd_col_start + i)
            if raw is not NOT_FOUND and raw is not None:
                cell.value        = raw * multiplier
                cell.number_format = number_format
                cell.alignment    = Alignment(horizontal='right')
            else:
                cell.value     = raw
                cell.alignment = Alignment(horizontal='left')

    # Auto-fit all columns (standard + USD)
    total_cols = len(OUTPUT_COLUMNS) + len(_NORMALIZABLE_COLS)
    for col_idx in range(1, total_cols + 1):
        if col_idx < usd_col_start:
            header_len = len(OUTPUT_COLUMNS[col_idx - 1])
        else:
            header_len = len(f'{_NORMALIZABLE_COLS[col_idx - usd_col_start]} (USD)')
        max_len = header_len
        for row_idx in range(2, len(all_results) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, min(len(str(cell_val)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    wb.save(output_path)
    logger.info(f"\nExcel output written to: {output_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def setup_logging(log_file: Optional[str]) -> logging.Logger:
    logger = logging.getLogger('parse_cafr')
    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s', datefmt='%H:%M:%S')

    # Always log to stderr
    sh = logging.StreamHandler(sys.stderr)
    sh.setLevel(logging.INFO)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    # Optionally log to file
    if log_file:
        fh = logging.FileHandler(log_file, encoding='utf-8')
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(fmt)
        logger.addHandler(fh)

    return logger


def collect_inputs(files: List[str], dir_path: Optional[str]) -> List[Path]:
    """Collect and deduplicate PDF input paths."""
    seen = set()
    paths = []
    for f in (files or []):
        p = Path(f).resolve()
        if p not in seen:
            seen.add(p)
            paths.append(p)
    if dir_path:
        d = Path(dir_path)
        for p in sorted(d.glob('*.pdf')):
            rp = p.resolve()
            if rp not in seen:
                seen.add(rp)
                paths.append(rp)
    return paths


def print_summary(all_results: List[Dict]):
    """Print a compact summary of extracted figures to stdout."""
    numeric_cols = [
        'Total Assets', 'Total Liabilities',
        'FB - Nonspendable', 'FB - Restricted', 'FB - Committed',
        'FB - Assigned', 'FB - Unassigned', 'Total Fund Balances',
        'Total Revenues', 'Total Expenditures', 'Total OFS (Uses)',
    ]
    print("\n" + "=" * 80)
    print("EXTRACTION RESULTS SUMMARY")
    print("=" * 80)
    for row in all_results:
        print(f"\n  {row['Source File']}")
        print(f"  Entity: {row['Entity Name']}  |  FY End: {row['Fiscal Year End']}  |  Units: {row['Reporting Units']}")
        for col in numeric_cols:
            val = row.get(col, NOT_FOUND)
            if val is NOT_FOUND:
                print(f"    {col:<30} NOT FOUND")
            else:
                print(f"    {col:<30} {val:>20,.2f}")
        if row.get('Extraction Notes'):
            print(f"  Notes: {row['Extraction Notes']}")
    print("\n" + "=" * 80)


def main():
    parser = argparse.ArgumentParser(
        description='Parse ACFR PDFs and extract General Fund figures to Excel.'
    )
    parser.add_argument(
        'files', nargs='*',
        help='One or more PDF file paths'
    )
    parser.add_argument(
        '--dir', metavar='DIRECTORY',
        help='Directory to scan for all .pdf files (non-recursive)'
    )
    parser.add_argument(
        '--output', default='cafr_output.xlsx', metavar='FILE',
        help='Output Excel file path (default: cafr_output.xlsx)'
    )
    parser.add_argument(
        '--log', metavar='FILE',
        help='Optional log file path (logging always goes to stderr too)'
    )
    args = parser.parse_args()

    logger = setup_logging(args.log)

    # Collect inputs
    input_paths = collect_inputs(args.files, args.dir)
    if not input_paths:
        logger.error("No input PDF files found. Provide file paths or --dir.")
        sys.exit(1)

    missing = [p for p in input_paths if not p.exists()]
    if missing:
        logger.error(f"Files not found: {[str(p) for p in missing]}")
        sys.exit(1)

    logger.info(f"Processing {len(input_paths)} PDF(s)")
    logger.info(f"Output: {args.output}")

    # Process each PDF
    all_results = []
    for pdf_path in input_paths:
        row = process_pdf(pdf_path, logger)
        all_results.append(row)

    # Print summary to stdout
    print_summary(all_results)

    # Write Excel
    output_path = Path(args.output)
    write_excel(all_results, output_path, logger)
    print(f"\nExcel written: {output_path.resolve()}")


if __name__ == '__main__':
    main()
